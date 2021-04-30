import os
import xlrd
import pickle
import openpyxl
import numpy as np
import urllib.request, json 
from flask import Flask, render_template
from flask import request, jsonify
from werkzeug.utils import secure_filename
from livereload import Server


app = Flask(__name__)
model = pickle.load(open('model.pkl', 'rb'))

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/page')
def page():
    return render_template('page.html')

@app.route('/result')
def result():
    return render_template('result.html')

@app.route('/gene')
def gene():
    return render_template('gene.html')

@app.route('/predict',methods=['GET', 'POST'])
def predict():
    '''
    For rendering results on HTML GUI
    '''
    if request.method == 'POST':
        f = request.files['file']
        f.filename = 'patient_data.xlsx'
        f.save(secure_filename(f.filename))
        # return 'file uploaded successfully'

    path = 'patient_data.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    # print(sheet_obj.cell(row = 2, column = 1).value)

    d1 = sheet_obj.cell(row = 2, column = 2).value
    d2 = sheet_obj.cell(row = 2, column = 3).value
    d3 = sheet_obj.cell(row = 2, column = 4).value
    d4 = sheet_obj.cell(row = 2, column = 5).value
    d5 = sheet_obj.cell(row = 2, column = 6).value
    d6 = sheet_obj.cell(row = 2, column = 7).value
    d7 = sheet_obj.cell(row = 2, column = 8).value
    d8 = sheet_obj.cell(row = 2, column = 9).value
    d9 = sheet_obj.cell(row = 2, column = 10).value
    d10= sheet_obj.cell(row = 2, column = 11).value
    d11= sheet_obj.cell(row = 2, column = 12).value
    d12= sheet_obj.cell(row = 2, column = 13).value
    d13= sheet_obj.cell(row = 2, column = 14).value
    d14= sheet_obj.cell(row = 2, column = 15).value
    d15= sheet_obj.cell(row = 2, column = 16).value
    d16= sheet_obj.cell(row = 2, column = 17).value
    d17= sheet_obj.cell(row = 2, column = 18).value
    d18= sheet_obj.cell(row = 2, column = 19).value
    d19= sheet_obj.cell(row = 2, column = 20).value
    d20= sheet_obj.cell(row = 2, column = 21).value
    arr=np.array([[d1,d2,d3,d4,d5,d6,d7,d8,d9,d10,d11,d12,d13,d14,d15,d16,d17,d18,d19,d20]])
    
    arr1=np.array([d1,d2,d3,d4,d5,d6,d7,d8,d9,d10,d11,d12,d13,d14,d15,d16,d17,d18,d19,d20])

    #int_features = [int(x) for x in request.form.values()]
    #final_features = [np.array(int_features)]
    prediction = model.predict(arr)

    feature_names = ['MCM2', 'DEK', 'H2AFZ', 'METRN', 'NFIB', 'TPX2', 'MCM6', 'TOPBP1', 'FEN1', 'TTK', 'ESR1', 'MELK', 'E2F3', 'CYP2B7P', 'TP53TG1', 'RARRES1', 'ROPN1B', 'PLCH1', 'SLC7A8', 'MLPH']
    from eli5 import show_prediction
    import eli5 as eli
    data = eli.show_prediction(model, arr1,
                        feature_names=feature_names,
                        show_feature_values=True).data
    file = open("static/eli5.html","w")
    file.write(data)

    output = prediction[0]
    if output==0:
        print("Chemotherapy Not Suitable")
    else:
        print("Chemotherapy Suitable")
    return render_template('result.html', data=data, prediction_text=output)

@app.route('/predict1',methods=['POST'])
def predict1():
    '''
    For rendering results on HTML GUI
    '''
    d1= request.form['a']
    d2= request.form['b']
    d3= request.form['c']
    d4= request.form['d']
    d5= request.form['e']
    d6= request.form['f']
    d7= request.form['g']
    d8= request.form['h']
    d9= request.form['i']
    d10= request.form['j']
    d11=request.form['k']
    d12= request.form['l']
    d13= request.form['m']
    d14= request.form['n']
    d15= request.form['o']
    d16= request.form['p']
    d17= request.form['q']
    d18= request.form['r']
    d19= request.form['s']
    d20= request.form['t']
    arr=np.array([[d1,d2,d3,d4,d5,d6,d7,d8,d9,d10,d11,d12,d13,d14,d15,d16,d17,d18,d19,d20]])

    arr1=np.array([d1,d2,d3,d4,d5,d6,d7,d8,d9,d10,d11,d12,d13,d14,d15,d16,d17,d18,d19,d20])

    #int_features = [int(x) for x in request.form.values()]
    #final_features = [np.array(int_features)]
    prediction1 = model.predict(arr)

    feature_names = ['MCM2', 'DEK', 'H2AFZ', 'METRN', 'NFIB', 'TPX2', 'MCM6', 'TOPBP1', 'FEN1', 'TTK', 'ESR1', 'MELK', 'E2F3', 'CYP2B7P', 'TP53TG1', 'RARRES1', 'ROPN1B', 'PLCH1', 'SLC7A8', 'MLPH']
    from eli5 import show_prediction
    import eli5 as eli
    data = eli.show_prediction(model, arr1,
                        feature_names=feature_names,
                        show_feature_values=True).data
    file = open("static/eli5.html","w")
    file.write(data)

    output = prediction1[0]
    if output==0:
        print("Chemotherapy Not Suitable")
    else:
        print("Chemotherapy Suitable")
    return render_template('result.html', data=prediction1, prediction_text=output)

@app.route('/search',methods=['POST'])
def search():
    if request.method == 'POST':
        word = request.form['query']
        search_word = word.upper()
        print(search_word)
        link = "https://clinicaltables.nlm.nih.gov/api/ncbi_genes/v3/search?terms="+search_word
        with urllib.request.urlopen(link) as url:
            data = json.loads(url.read().decode())
            # print(data)
    return jsonify({ 'htmlresponse': render_template('table.html', data = data, n = int(len(data[1])), word = search_word ) })

if __name__ == "__main__":
    # app.run(debug=True)
    server = Server(app.wsgi_app)
    server.serve()